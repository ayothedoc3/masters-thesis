"""
Export scored spreadsheet data to CSV files for R analysis.
Produces:
  - data/csv/master_dataset.csv (updated with all scores, 200 rows)
  - data/csv/second_rater_scores.csv (40 rows for IRR)
"""

import openpyxl
import csv
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET_PATH = os.path.join(BASE_DIR, "data_collection_spreadsheet.xlsx")
MASTER_CSV_PATH = os.path.join(BASE_DIR, "csv", "master_dataset.csv")
RATER2_CSV_PATH = os.path.join(BASE_DIR, "csv", "second_rater_scores.csv")

# Column mapping: column_index -> CSV header name
COLUMNS = {
    1: "Item_ID",
    2: "Platform",
    3: "Search_Term",
    4: "URL",
    5: "Title_Caption",
    6: "Creator_Name",
    7: "Creator_Credentials",
    8: "Date_Posted",
    9: "Views",
    10: "Likes",
    11: "Shares",
    12: "Comments",
    13: "Creator_Type",
    14: "Content_Format",
}

# DISCERN Q1-Q16 in columns O(15) through AD(30)
for q in range(1, 17):
    COLUMNS[14 + q] = f"DISCERN_Q{q}"

# Calculated columns
COLUMNS[31] = "DISCERN_Section1"
COLUMNS[32] = "DISCERN_Section2"
COLUMNS[33] = "DISCERN_Total"
COLUMNS[34] = "DISCERN_Category"

# JAMA
COLUMNS[35] = "JAMA_Authorship"
COLUMNS[36] = "JAMA_Attribution"
COLUMNS[37] = "JAMA_Disclosure"
COLUMNS[38] = "JAMA_Currency"
COLUMNS[39] = "JAMA_Total"

# Notes
COLUMNS[40] = "Notes"


def export_sheet(ws, output_path, sheet_name):
    """Export a worksheet to CSV, computing formula values manually."""
    headers = [COLUMNS[i] for i in sorted(COLUMNS.keys())]
    rows_written = 0

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for row_num in range(2, ws.max_row + 1):
            item_id = ws.cell(row=row_num, column=1).value
            if not item_id:
                continue

            row_data = []
            for col_idx in sorted(COLUMNS.keys()):
                cell = ws.cell(row=row_num, column=col_idx)
                value = cell.value

                # For formula cells, compute the value manually
                if isinstance(value, str) and value.startswith("="):
                    col_name = COLUMNS[col_idx]

                    if col_name == "DISCERN_Section1":
                        # SUM of Q1-Q8 (cols 15-22)
                        vals = [ws.cell(row=row_num, column=c).value for c in range(15, 23)]
                        value = sum(v for v in vals if isinstance(v, (int, float)))

                    elif col_name == "DISCERN_Section2":
                        # SUM of Q9-Q15 (cols 23-29)
                        vals = [ws.cell(row=row_num, column=c).value for c in range(23, 30)]
                        value = sum(v for v in vals if isinstance(v, (int, float)))

                    elif col_name == "DISCERN_Total":
                        # SUM of Q1-Q16 (cols 15-30)
                        vals = [ws.cell(row=row_num, column=c).value for c in range(15, 31)]
                        value = sum(v for v in vals if isinstance(v, (int, float)))

                    elif col_name == "DISCERN_Category":
                        total = sum(
                            v for v in [ws.cell(row=row_num, column=c).value for c in range(15, 31)]
                            if isinstance(v, (int, float))
                        )
                        if total <= 26:
                            value = "Very Poor"
                        elif total <= 38:
                            value = "Poor"
                        elif total <= 50:
                            value = "Fair"
                        elif total <= 62:
                            value = "Good"
                        else:
                            value = "Excellent"

                    elif col_name == "JAMA_Total":
                        jama_vals = [ws.cell(row=row_num, column=c).value for c in range(35, 39)]
                        value = sum(1 for v in jama_vals if v == "Y")

                row_data.append(value if value is not None else "")

            writer.writerow(row_data)
            rows_written += 1

    print(f"  {sheet_name}: {rows_written} rows -> {output_path}")
    return rows_written


def main():
    print("=" * 60)
    print("Exporting scored data to CSV")
    print("=" * 60)

    wb = openpyxl.load_workbook(SPREADSHEET_PATH)

    # Export main data
    print("\nExporting Data Collection sheet...")
    n1 = export_sheet(wb["Data Collection"], MASTER_CSV_PATH, "Data Collection")

    # Export second rater
    print("\nExporting Second Rater sheet...")
    n2 = export_sheet(wb["Second Rater"], RATER2_CSV_PATH, "Second Rater")

    # Quick verification
    print("\nVerification:")
    import csv as csv_mod
    for path, label in [(MASTER_CSV_PATH, "Master"), (RATER2_CSV_PATH, "Second Rater")]:
        with open(path, "r", encoding="utf-8") as f:
            reader = csv_mod.DictReader(f)
            rows = list(reader)
            print(f"  {label}: {len(rows)} rows, {len(rows[0])} columns")
            # Check a row
            r = rows[0]
            print(f"    First row: {r['Item_ID']}, DISCERN_Total={r['DISCERN_Total']}, "
                  f"DISCERN_Category={r['DISCERN_Category']}, JAMA_Total={r['JAMA_Total']}, "
                  f"Creator_Type={r['Creator_Type']}")

    print("\n" + "=" * 60)
    print("SUCCESS: CSVs exported for R analysis!")
    print("=" * 60)


if __name__ == "__main__":
    main()
