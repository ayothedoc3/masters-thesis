"""
Import master_dataset.csv into the existing data_collection_spreadsheet.xlsx.
Matches rows by Item_ID and fills columns D-N with collected data.
"""

import openpyxl
import csv
import os
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET_PATH = os.path.join(BASE_DIR, "data_collection_spreadsheet.xlsx")
MASTER_CSV_PATH = os.path.join(BASE_DIR, "csv", "master_dataset.csv")

# Column mapping: spreadsheet column index -> CSV field name
# A(1)=Item_ID, B(2)=Platform, C(3)=Search_Term (already filled)
# D(4)=URL, E(5)=Title_Caption, F(6)=Creator_Name, G(7)=Creator_Credentials,
# H(8)=Date_Posted, I(9)=Views, J(10)=Likes, K(11)=Shares, L(12)=Comments,
# M(13)=Creator_Type, N(14)=Content_Format
IMPORT_COLUMNS = {
    4: "URL",
    5: "Title_Caption",
    6: "Creator_Name",
    7: "Creator_Credentials",
    8: "Date_Posted",
    9: "Views",
    10: "Likes",
    11: "Shares",
    12: "Comments",
    13: "Creator_Type",
    14: "Content_Format",
}


def parse_date(date_str):
    """Try to parse various date formats into a clean date string."""
    if not date_str or date_str.strip() == "":
        return ""
    date_str = date_str.strip()
    # Try ISO format (from Apify)
    for fmt in ["%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S.000Z",
                "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # If it's a Unix timestamp (seconds)
    try:
        ts = int(date_str)
        if ts > 1000000000:  # looks like a Unix timestamp
            dt = datetime.fromtimestamp(ts)
            return dt.strftime("%Y-%m-%d")
    except (ValueError, OSError):
        pass
    return date_str  # Return as-is if we can't parse


def parse_number(val):
    """Convert string numbers to integers, handling K/M suffixes."""
    if not val or val.strip() == "":
        return None
    val = val.strip().replace(",", "")
    # Handle -1 (Instagram sometimes returns this)
    if val == "-1":
        return None
    try:
        return int(float(val))
    except ValueError:
        return None


def read_csv(csv_path):
    """Read master CSV and return dict keyed by Item_ID."""
    data = {}
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            item_id = row.get("Item_ID", "").strip()
            if item_id:
                data[item_id] = row
    return data


def main():
    print("=" * 60)
    print("Importing CSV data into spreadsheet")
    print("=" * 60)

    # Read CSV
    csv_data = read_csv(MASTER_CSV_PATH)
    print(f"  Read {len(csv_data)} rows from master_dataset.csv")

    # Open spreadsheet
    wb = openpyxl.load_workbook(SPREADSHEET_PATH)
    ws = wb["Data Collection"]
    print(f"  Opened spreadsheet: {ws.max_row} rows x {ws.max_column} columns")

    # Import data
    imported = 0
    skipped = 0

    for row_num in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row_num, column=1).value
        if not item_id:
            continue

        csv_row = csv_data.get(item_id)
        if not csv_row:
            print(f"  WARNING: {item_id} not found in CSV")
            skipped += 1
            continue

        for col_idx, csv_field in IMPORT_COLUMNS.items():
            raw_value = csv_row.get(csv_field, "")

            if csv_field == "Date_Posted":
                value = parse_date(raw_value)
            elif csv_field in ("Views", "Likes", "Shares", "Comments"):
                value = parse_number(raw_value)
            elif csv_field == "Title_Caption":
                # Truncate very long captions for spreadsheet readability
                value = raw_value[:500] if raw_value else ""
            else:
                value = raw_value.strip() if raw_value else ""

            if value is not None and value != "":
                ws.cell(row=row_num, column=col_idx, value=value)

        imported += 1

    # Also update the Second Rater sheet with URLs so the rater can access content
    if "Second Rater" in wb.sheetnames:
        ws2 = wb["Second Rater"]
        rater_imported = 0
        for row_num in range(2, ws2.max_row + 1):
            item_id = ws2.cell(row=row_num, column=1).value
            if not item_id:
                continue
            csv_row = csv_data.get(item_id)
            if not csv_row:
                continue
            for col_idx, csv_field in IMPORT_COLUMNS.items():
                raw_value = csv_row.get(csv_field, "")
                if csv_field == "Date_Posted":
                    value = parse_date(raw_value)
                elif csv_field in ("Views", "Likes", "Shares", "Comments"):
                    value = parse_number(raw_value)
                elif csv_field == "Title_Caption":
                    value = raw_value[:500] if raw_value else ""
                else:
                    value = raw_value.strip() if raw_value else ""
                if value is not None and value != "":
                    ws2.cell(row=row_num, column=col_idx, value=value)
            rater_imported += 1
        print(f"  Second Rater sheet: {rater_imported} rows updated")

    # Save
    wb.save(SPREADSHEET_PATH)
    print()
    print(f"  Imported: {imported} rows")
    print(f"  Skipped: {skipped} rows")
    print(f"  Saved to: {SPREADSHEET_PATH}")

    # Quick verification
    wb2 = openpyxl.load_workbook(SPREADSHEET_PATH)
    ws2 = wb2["Data Collection"]
    print()
    print("  Spot checks:")
    print(f"    WEB-01 URL: {ws2['D2'].value}")
    print(f"    WEB-01 Title: {ws2['E2'].value}")
    print(f"    YT-01 URL: {ws2['D52'].value}")
    print(f"    YT-01 Creator: {ws2['F52'].value}")
    print(f"    TT-01 URL: {ws2['D102'].value}")
    print(f"    IG-01 URL: {ws2['D152'].value}")
    print()
    print("=" * 60)
    print("SUCCESS: Data imported into spreadsheet!")
    print("=" * 60)


if __name__ == "__main__":
    main()
