"""
Build comprehensive data collection Excel spreadsheet for Master's thesis.
Requires: pip install openpyxl
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import (
    CellIsRule, ColorScaleRule, FormulaRule
)
import random
import os

# =============================================================================
# Configuration
# =============================================================================

OUTPUT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "data_collection_spreadsheet.xlsx"
)

PLATFORMS = [
    {"prefix": "WEB", "name": "Website", "count": 50},
    {"prefix": "YT",  "name": "YouTube", "count": 50},
    {"prefix": "TT",  "name": "TikTok",  "count": 50},
    {"prefix": "IG",  "name": "Instagram", "count": 50},
]

SEARCH_TERMS = [
    "how to start exercising",
    "best exercises to lose weight",
    "strength training for beginners",
    "physical activity guidelines",
    "home workout routine",
]

# 10 items per search term per platform
ITEMS_PER_TERM = 10

# Colors
HEADER_BG = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"

# Alternating platform row colors
PLATFORM_COLORS = {
    "Website":   "D6EAF8",   # light blue
    "YouTube":   "F5B7B1",   # light red
    "TikTok":    "D5F5E3",   # light green
    "Instagram": "E8DAEF",   # light purple
}

# Conditional formatting colors for DISCERN_Category
CATEGORY_COLORS = {
    "Very Poor": "FF0000",   # red
    "Poor":      "FF8C00",   # orange
    "Fair":      "FFD700",   # yellow
    "Good":      "90EE90",   # light green
    "Excellent": "228B22",   # dark green
}

# Dropdown options
CREATOR_TYPES = [
    "Healthcare professional",
    "Certified fitness professional",
    "Fitness influencer",
    "General user",
    "Organization",
]

CONTENT_FORMATS = [
    "Text article",
    "Long video (>5min)",
    "Short video (<60s)",
    "Image+caption",
    "Carousel",
]

DISCERN_SCALE = "1,2,3,4,5"
YN_OPTIONS = "Y,N"

# =============================================================================
# Column definitions for Sheet 1: Data Collection
# =============================================================================

# Column letter -> (Header, width, number_format or None)
# We'll build this programmatically

SHEET1_COLUMNS = []

# A: Item_ID
SHEET1_COLUMNS.append(("Item_ID", 12, None))
# B: Platform
SHEET1_COLUMNS.append(("Platform", 14, None))
# C: Search_Term
SHEET1_COLUMNS.append(("Search_Term", 32, None))
# D: URL
SHEET1_COLUMNS.append(("URL", 45, None))
# E: Title_Caption
SHEET1_COLUMNS.append(("Title_Caption", 30, None))
# F: Creator_Name
SHEET1_COLUMNS.append(("Creator_Name", 18, None))
# G: Creator_Credentials
SHEET1_COLUMNS.append(("Creator_Credentials", 22, None))
# H: Date_Posted
SHEET1_COLUMNS.append(("Date_Posted", 14, "YYYY-MM-DD"))
# I: Views
SHEET1_COLUMNS.append(("Views", 12, "#,##0"))
# J: Likes
SHEET1_COLUMNS.append(("Likes", 12, "#,##0"))
# K: Shares
SHEET1_COLUMNS.append(("Shares", 12, "#,##0"))
# L: Comments
SHEET1_COLUMNS.append(("Comments", 12, "#,##0"))
# M: Creator_Type (dropdown)
SHEET1_COLUMNS.append(("Creator_Type", 28, None))
# N: Content_Format (dropdown)
SHEET1_COLUMNS.append(("Content_Format", 22, None))

# O-AD: DISCERN Q1-Q16 (columns 15-30)
for q in range(1, 17):
    SHEET1_COLUMNS.append((f"DISCERN_Q{q}", 12, None))

# AE: DISCERN_Section1 (formula)
SHEET1_COLUMNS.append(("DISCERN_Section1", 16, "#,##0"))
# AF: DISCERN_Section2 (formula)
SHEET1_COLUMNS.append(("DISCERN_Section2", 16, "#,##0"))
# AG: DISCERN_Total (formula)
SHEET1_COLUMNS.append(("DISCERN_Total", 14, "#,##0"))
# AH: DISCERN_Category (formula)
SHEET1_COLUMNS.append(("DISCERN_Category", 18, None))
# AI: JAMA_Authorship (dropdown)
SHEET1_COLUMNS.append(("JAMA_Authorship", 16, None))
# AJ: JAMA_Attribution (dropdown)
SHEET1_COLUMNS.append(("JAMA_Attribution", 16, None))
# AK: JAMA_Disclosure (dropdown)
SHEET1_COLUMNS.append(("JAMA_Disclosure", 16, None))
# AL: JAMA_Currency (dropdown)
SHEET1_COLUMNS.append(("JAMA_Currency", 15, None))
# AM: JAMA_Total (formula)
SHEET1_COLUMNS.append(("JAMA_Total", 12, "#,##0"))
# AN: Notes
SHEET1_COLUMNS.append(("Notes", 35, None))

# Total columns: 40 (A through AN)

# =============================================================================
# Helper functions
# =============================================================================

def generate_items():
    """Generate all 200 item rows with pre-filled data."""
    rows = []
    for platform in PLATFORMS:
        for i in range(1, platform["count"] + 1):
            term_index = (i - 1) // ITEMS_PER_TERM
            item_id = f"{platform['prefix']}-{i:02d}"
            rows.append({
                "item_id": item_id,
                "platform": platform["name"],
                "search_term": SEARCH_TERMS[term_index],
            })
    return rows


def apply_header_style(ws, num_cols):
    """Apply styling to the header row."""
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11, name="Calibri")
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set row height for header
    ws.row_dimensions[1].height = 35


def apply_column_widths(ws, columns):
    """Set column widths."""
    for i, (header, width, fmt) in enumerate(columns, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width


def apply_row_colors(ws, items, start_row, num_cols):
    """Apply alternating platform background colors."""
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    cell_alignment = Alignment(vertical="center", wrap_text=False)

    for idx, item in enumerate(items):
        row = start_row + idx
        color = PLATFORM_COLORS.get(item["platform"], "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = cell_alignment


def add_data_validations_sheet1(ws, start_row, end_row):
    """Add dropdown validations to Sheet 1."""
    # Creator_Type - column M (13)
    dv_creator = DataValidation(
        type="list",
        formula1='"' + ",".join(CREATOR_TYPES) + '"',
        allow_blank=True
    )
    dv_creator.error = "Please select a valid Creator Type"
    dv_creator.errorTitle = "Invalid Creator Type"
    dv_creator.prompt = "Select the creator type"
    dv_creator.promptTitle = "Creator Type"
    dv_creator.showErrorMessage = True
    dv_creator.showInputMessage = True
    col_m = get_column_letter(13)
    dv_creator.add(f"{col_m}{start_row}:{col_m}{end_row}")
    ws.add_data_validation(dv_creator)

    # Content_Format - column N (14)
    dv_format = DataValidation(
        type="list",
        formula1='"' + ",".join(CONTENT_FORMATS) + '"',
        allow_blank=True
    )
    dv_format.error = "Please select a valid Content Format"
    dv_format.errorTitle = "Invalid Content Format"
    dv_format.prompt = "Select the content format"
    dv_format.promptTitle = "Content Format"
    dv_format.showErrorMessage = True
    dv_format.showInputMessage = True
    col_n = get_column_letter(14)
    dv_format.add(f"{col_n}{start_row}:{col_n}{end_row}")
    ws.add_data_validation(dv_format)

    # DISCERN Q1-Q16 - columns O through AD (15-30)
    dv_discern = DataValidation(
        type="list",
        formula1='"' + DISCERN_SCALE + '"',
        allow_blank=True
    )
    dv_discern.error = "Please enter a value between 1 and 5"
    dv_discern.errorTitle = "Invalid DISCERN Score"
    dv_discern.prompt = "Rate 1 (No) to 5 (Yes)"
    dv_discern.promptTitle = "DISCERN Rating"
    dv_discern.showErrorMessage = True
    dv_discern.showInputMessage = True
    for q_col in range(15, 31):  # O=15 through AD=30
        col_letter = get_column_letter(q_col)
        dv_discern.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    ws.add_data_validation(dv_discern)

    # JAMA Y/N - columns AI through AL (35-38)
    dv_yn = DataValidation(
        type="list",
        formula1='"' + YN_OPTIONS + '"',
        allow_blank=True
    )
    dv_yn.error = "Please enter Y or N"
    dv_yn.errorTitle = "Invalid Entry"
    dv_yn.prompt = "Enter Y (Yes) or N (No)"
    dv_yn.promptTitle = "JAMA Benchmark"
    dv_yn.showErrorMessage = True
    dv_yn.showInputMessage = True
    for jama_col in range(35, 39):  # AI=35, AJ=36, AK=37, AL=38
        col_letter = get_column_letter(jama_col)
        dv_yn.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    ws.add_data_validation(dv_yn)


def add_formulas_sheet1(ws, start_row, end_row):
    """Add formulas to calculated columns in Sheet 1."""
    for row in range(start_row, end_row + 1):
        # DISCERN Q1-Q8 are in columns O(15) through V(22)
        # DISCERN Q9-Q15 are in columns W(23) through AC(29)
        # DISCERN Q16 is in column AD(30)

        # Column letters for Q1-Q8
        q1_col = get_column_letter(15)  # O
        q8_col = get_column_letter(22)  # V

        # Column letters for Q9-Q15
        q9_col = get_column_letter(23)  # W
        q15_col = get_column_letter(29)  # AC

        # Column letters for Q1-Q16
        q16_col = get_column_letter(30)  # AD

        # AE (31): DISCERN_Section1 = SUM(Q1:Q8)
        ws.cell(row=row, column=31).value = f"=SUM({q1_col}{row}:{q8_col}{row})"

        # AF (32): DISCERN_Section2 = SUM(Q9:Q15)
        ws.cell(row=row, column=32).value = f"=SUM({q9_col}{row}:{q15_col}{row})"

        # AG (33): DISCERN_Total = SUM(Q1:Q16)
        ws.cell(row=row, column=33).value = f"=SUM({q1_col}{row}:{q16_col}{row})"

        # AH (34): DISCERN_Category based on total
        total_ref = f"{get_column_letter(33)}{row}"
        formula = (
            f'=IF({total_ref}="","",IF({total_ref}<=26,"Very Poor",'
            f'IF({total_ref}<=38,"Poor",IF({total_ref}<=50,"Fair",'
            f'IF({total_ref}<=62,"Good","Excellent")))))'
        )
        ws.cell(row=row, column=34).value = formula

        # AM (39): JAMA_Total = COUNTIF of Y values in AI:AL
        ai_col = get_column_letter(35)
        al_col = get_column_letter(38)
        ws.cell(row=row, column=39).value = (
            f'=COUNTIF({ai_col}{row}:{al_col}{row},"Y")'
        )


def add_number_formats_sheet1(ws, columns, start_row, end_row):
    """Apply number formats to data cells."""
    for col_idx, (header, width, fmt) in enumerate(columns, 1):
        if fmt is not None:
            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=col_idx).number_format = fmt


def add_conditional_formatting_sheet1(ws, start_row, end_row):
    """Add conditional formatting rules."""
    # DISCERN_Category column AH (34)
    cat_col = get_column_letter(34)
    cat_range = f"{cat_col}{start_row}:{cat_col}{end_row}"

    # Very Poor = red
    ws.conditional_formatting.add(
        cat_range,
        CellIsRule(
            operator="equal",
            formula=['"Very Poor"'],
            fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            font=Font(color="FFFFFF", bold=True),
        )
    )
    # Poor = orange
    ws.conditional_formatting.add(
        cat_range,
        CellIsRule(
            operator="equal",
            formula=['"Poor"'],
            fill=PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
            font=Font(color="FFFFFF", bold=True),
        )
    )
    # Fair = yellow
    ws.conditional_formatting.add(
        cat_range,
        CellIsRule(
            operator="equal",
            formula=['"Fair"'],
            fill=PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            font=Font(bold=True),
        )
    )
    # Good = light green
    ws.conditional_formatting.add(
        cat_range,
        CellIsRule(
            operator="equal",
            formula=['"Good"'],
            fill=PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            font=Font(bold=True),
        )
    )
    # Excellent = dark green
    ws.conditional_formatting.add(
        cat_range,
        CellIsRule(
            operator="equal",
            formula=['"Excellent"'],
            fill=PatternFill(start_color="228B22", end_color="228B22", fill_type="solid"),
            font=Font(color="FFFFFF", bold=True),
        )
    )

    # DISCERN_Total column AG (33) - color scale red-yellow-green
    total_col = get_column_letter(33)
    total_range = f"{total_col}{start_row}:{total_col}{end_row}"

    ws.conditional_formatting.add(
        total_range,
        ColorScaleRule(
            start_type="num", start_value=16, start_color="FF0000",
            mid_type="num", mid_value=48, mid_color="FFD700",
            end_type="num", end_value=80, end_color="228B22",
        )
    )

    # JAMA_Total column AM (39) - color scale
    jama_col = get_column_letter(39)
    jama_range = f"{jama_col}{start_row}:{jama_col}{end_row}"

    ws.conditional_formatting.add(
        jama_range,
        ColorScaleRule(
            start_type="num", start_value=0, start_color="FF0000",
            mid_type="num", mid_value=2, mid_color="FFD700",
            end_type="num", end_value=4, end_color="228B22",
        )
    )


# =============================================================================
# Build Sheet 1: Data Collection
# =============================================================================

def build_sheet1(wb):
    """Build the main Data Collection sheet."""
    ws = wb.active
    ws.title = "Data Collection"

    num_cols = len(SHEET1_COLUMNS)
    items = generate_items()
    start_row = 2
    end_row = start_row + len(items) - 1  # 201

    # --- Write headers ---
    for col_idx, (header, width, fmt) in enumerate(SHEET1_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # --- Apply header styling ---
    apply_header_style(ws, num_cols)

    # --- Set column widths ---
    apply_column_widths(ws, SHEET1_COLUMNS)

    # --- Pre-fill data rows ---
    for idx, item in enumerate(items):
        row = start_row + idx
        ws.cell(row=row, column=1, value=item["item_id"])    # A: Item_ID
        ws.cell(row=row, column=2, value=item["platform"])    # B: Platform
        ws.cell(row=row, column=3, value=item["search_term"]) # C: Search_Term

    # --- Apply row colors ---
    apply_row_colors(ws, items, start_row, num_cols)

    # --- Add data validations ---
    add_data_validations_sheet1(ws, start_row, end_row)

    # --- Add formulas ---
    add_formulas_sheet1(ws, start_row, end_row)

    # --- Add number formats ---
    add_number_formats_sheet1(ws, SHEET1_COLUMNS, start_row, end_row)

    # --- Add conditional formatting ---
    add_conditional_formatting_sheet1(ws, start_row, end_row)

    # --- Freeze panes (freeze header row) ---
    ws.freeze_panes = "A2"

    # --- Set print area and page setup ---
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
        fitToPage=True
    )

    print(f"  Sheet 1 'Data Collection': {len(items)} rows, {num_cols} columns")
    return items  # Return for use in Sheet 3


# =============================================================================
# Build Sheet 2: Summary Statistics
# =============================================================================

def build_sheet2(wb):
    """Build the Summary Statistics placeholder sheet."""
    ws = wb.create_sheet("Summary Statistics")

    headers = [
        "Platform",
        "N",
        "DISCERN_Mean",
        "DISCERN_SD",
        "DISCERN_Median",
        "DISCERN_IQR",
        "DISCERN_Min",
        "DISCERN_Max",
        "Pct_Very_Poor",
        "Pct_Poor",
        "Pct_Fair",
        "Pct_Good",
        "Pct_Excellent",
        "JAMA_Mean",
        "JAMA_SD",
        "JAMA_Median",
        "JAMA_Authorship_Pct",
        "JAMA_Attribution_Pct",
        "JAMA_Disclosure_Pct",
        "JAMA_Currency_Pct",
    ]

    num_cols = len(headers)

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Apply header styling
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11, name="Calibri")
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 35

    # Pre-fill platform names and "Overall" row
    platform_names = ["Website", "YouTube", "TikTok", "Instagram", "Overall"]
    for idx, name in enumerate(platform_names):
        row = 2 + idx
        ws.cell(row=row, column=1, value=name)
        fill_color = PLATFORM_COLORS.get(name, "F2F2F2")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)

    # Add note
    ws.cell(row=8, column=1, value="Note: This sheet will be populated by R statistical analysis.")
    ws.cell(row=8, column=1).font = Font(italic=True, color="808080")

    # Freeze panes
    ws.freeze_panes = "A2"

    print(f"  Sheet 2 'Summary Statistics': {len(headers)} columns, placeholder rows for 4 platforms + Overall")


# =============================================================================
# Build Sheet 3: Second Rater
# =============================================================================

def build_sheet3(wb, all_items):
    """Build the Second Rater sheet with 40 randomly selected rows."""
    ws = wb.create_sheet("Second Rater")

    num_cols = len(SHEET1_COLUMNS)

    # Select 10 random items per platform (for inter-rater reliability)
    random.seed(42)  # For reproducibility
    selected_items = []
    for platform in PLATFORMS:
        platform_items = [item for item in all_items if item["platform"] == platform["name"]]
        selected = random.sample(platform_items, 10)
        # Sort by item_id for clean presentation
        selected.sort(key=lambda x: x["item_id"])
        selected_items.extend(selected)

    start_row = 2
    end_row = start_row + len(selected_items) - 1  # 41

    # --- Write headers (same as Sheet 1) ---
    for col_idx, (header, width, fmt) in enumerate(SHEET1_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # --- Apply header styling ---
    apply_header_style(ws, num_cols)

    # --- Set column widths ---
    apply_column_widths(ws, SHEET1_COLUMNS)

    # --- Pre-fill data rows ---
    for idx, item in enumerate(selected_items):
        row = start_row + idx
        ws.cell(row=row, column=1, value=item["item_id"])     # A: Item_ID
        ws.cell(row=row, column=2, value=item["platform"])     # B: Platform
        ws.cell(row=row, column=3, value=item["search_term"])  # C: Search_Term

    # --- Apply row colors ---
    apply_row_colors(ws, selected_items, start_row, num_cols)

    # --- Add data validations (same as Sheet 1) ---
    add_data_validations_sheet1(ws, start_row, end_row)

    # --- Add formulas (same as Sheet 1) ---
    add_formulas_sheet1(ws, start_row, end_row)

    # --- Add number formats ---
    add_number_formats_sheet1(ws, SHEET1_COLUMNS, start_row, end_row)

    # --- Add conditional formatting ---
    add_conditional_formatting_sheet1(ws, start_row, end_row)

    # --- Freeze panes ---
    ws.freeze_panes = "A2"

    print(f"  Sheet 3 'Second Rater': {len(selected_items)} rows (10 per platform), {num_cols} columns")
    print(f"  Selected items for second rating:")
    for platform in PLATFORMS:
        platform_ids = [item["item_id"] for item in selected_items if item["platform"] == platform["name"]]
        print(f"    {platform['name']}: {', '.join(platform_ids)}")


# =============================================================================
# Main
# =============================================================================

def main():
    print("=" * 70)
    print("Building Data Collection Spreadsheet for Master's Thesis")
    print("=" * 70)
    print()

    wb = Workbook()

    # Build all sheets
    print("Building sheets...")
    all_items = build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb, all_items)

    # Save
    print()
    print(f"Saving to: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print()

    # Verification
    print("=" * 70)
    print("VERIFICATION")
    print("=" * 70)

    # Reload and verify
    wb2 = openpyxl.load_workbook(OUTPUT_PATH)
    print(f"  Sheets: {wb2.sheetnames}")
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"  '{name}': {ws.max_row} rows x {ws.max_column} columns")

    # Verify Sheet 1 content
    ws1 = wb2["Data Collection"]
    print()
    print("  Sheet 1 spot checks:")
    print(f"    A1 (header): {ws1['A1'].value}")
    print(f"    A2 (first item): {ws1['A2'].value}")
    print(f"    B2 (first platform): {ws1['B2'].value}")
    print(f"    C2 (first term): {ws1['C2'].value}")
    print(f"    A51 (50th item): {ws1['A51'].value}")
    print(f"    A52 (51st item): {ws1['A52'].value}")
    print(f"    B52 (YT platform): {ws1['B52'].value}")
    print(f"    A201 (last item): {ws1['A201'].value}")
    print(f"    B201 (last platform): {ws1['B201'].value}")
    print(f"    C201 (last term): {ws1['C201'].value}")

    # Verify formulas
    print()
    print("  Formula checks (row 2):")
    print(f"    AE2 (Section1): {ws1['AE2'].value}")
    print(f"    AF2 (Section2): {ws1['AF2'].value}")
    print(f"    AG2 (Total): {ws1['AG2'].value}")
    print(f"    AH2 (Category): {ws1['AH2'].value}")
    print(f"    AM2 (JAMA_Total): {ws1['AM2'].value}")

    # Verify data validations
    print()
    print(f"  Data validations in Sheet 1: {len(ws1.data_validations.dataValidation)}")
    print(f"  Conditional formatting rules in Sheet 1: {len(ws1.conditional_formatting._cf_rules)}")

    file_size = os.path.getsize(OUTPUT_PATH)
    print()
    print(f"  File size: {file_size:,} bytes ({file_size/1024:.1f} KB)")
    print()
    print("=" * 70)
    print("SUCCESS: Spreadsheet created successfully!")
    print("=" * 70)


if __name__ == "__main__":
    main()
