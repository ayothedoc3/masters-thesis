"""
Add a 'Scoring Guide' sheet as the first tab in the spreadsheet
with step-by-step instructions for the manual scoring phase.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET_PATH = os.path.join(BASE_DIR, "data_collection_spreadsheet.xlsx")

HEADER_BG = "1F4E79"
SECTION_BG = "2E75B6"
LIGHT_BG = "D6EAF8"
TIP_BG = "FFF9C4"
WHITE = "FFFFFF"


def main():
    wb = openpyxl.load_workbook(SPREADSHEET_PATH)

    # Create the sheet and move it to position 0
    ws = wb.create_sheet("Scoring Guide", 0)

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 85

    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    title_font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    section_font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    bold_font = Font(bold=True, size=11, name="Calibri")
    normal_font = Font(size=11, name="Calibri")
    tip_font = Font(size=11, italic=True, name="Calibri", color="7F6000")
    small_font = Font(size=10, name="Calibri", color="555555")

    title_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    section_fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
    light_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    tip_fill = PatternFill(start_color=TIP_BG, end_color=TIP_BG, fill_type="solid")

    wrap = Alignment(wrap_text=True, vertical="top")
    center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

    row = 1

    def write_title(text):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = center_wrap
        ws.row_dimensions[row].height = 40
        row += 1

    def write_section(text):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 30
        row += 1

    def write_line(text, font=normal_font, fill=None, height=None):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font
        cell.alignment = wrap
        if fill:
            cell.fill = fill
            ws.cell(row=row, column=2).fill = fill
        if height:
            ws.row_dimensions[row].height = height
        row += 1

    def write_blank():
        nonlocal row
        ws.row_dimensions[row].height = 8
        row += 1

    # =========================================================================
    # CONTENT
    # =========================================================================

    write_title("SCORING GUIDE — Manual Assessment Phase")
    write_blank()

    # --- SETUP ---
    write_section("SETUP (do once before starting)")
    write_line("1. Keep this spreadsheet open — score directly in the 'Data Collection' tab", bold_font)
    write_line("2. Open the DISCERN quick reference card:  scoring-guides/discern_quick_reference.md")
    write_line("3. Have the screenshots folder open for quick visual reference:  data/screenshots/")
    write_line("4. Save frequently as you work (Ctrl+S)")
    write_blank()

    # --- PER ITEM WORKFLOW ---
    write_section("PER-ITEM WORKFLOW (repeat for each of the 200 rows)")
    write_blank()

    write_line("STEP 1 — Open the content", bold_font, light_fill)
    write_line("  Click the URL in column D to open the content in your browser")
    write_line("  Websites: read the full article")
    write_line("  YouTube: watch the video (at least skim key sections)")
    write_line("  TikTok / Instagram: watch the video or read caption + visuals")
    write_blank()

    write_line("STEP 2 — Fill metadata columns G, M, N", bold_font, light_fill)
    write_line("  Column G (Creator_Credentials): Check their bio/about page for qualifications")
    write_line('     Examples: "MD", "NASM-CPT", "PhD in Exercise Science", "RN"')
    write_line("  Column M (Creator_Type): Select from dropdown:")
    write_line("     - Healthcare professional  (doctor, nurse, physiotherapist)")
    write_line("     - Certified fitness professional  (personal trainer, CSCS, NASM-CPT)")
    write_line("     - Fitness influencer  (large following, no formal credentials)")
    write_line("     - General user  (regular person sharing personal experience)")
    write_line("     - Organization  (CDC, NHS, WHO, Mayo Clinic, gym chain)")
    write_line("  Column N (Content_Format): Verify/correct the pre-filled value")
    write_blank()

    write_line("STEP 3 — Score DISCERN Q1-Q16  (columns O through AD)", bold_font, light_fill)
    write_line("  Rate each question from 1 (definite No) to 5 (definite Yes)")
    write_line('  For PA content, "treatment" = the exercise or physical activity being recommended')
    write_blank()

    write_line("  Q1   Are the aims of the publication clear?", bold_font)
    write_line("  Q2   Does it achieve its aims?", bold_font)
    write_line("  Q3   Is it relevant to the reader?", bold_font)
    write_line("  Q4   Are the sources of information clear? (references, citations)", bold_font)
    write_line("  Q5   Is the information current / is it dated?", bold_font)
    write_line("  Q6   Is it balanced and unbiased?", bold_font)
    write_line("  Q7   Does it list additional sources of support / further reading?", bold_font)
    write_line("  Q8   Does it mention areas of uncertainty?", bold_font)
    write_line("  Q9   Does it describe how treatments/exercises work?", bold_font)
    write_line("  Q10  Does it describe the benefits of the treatment/exercise?", bold_font)
    write_line("  Q11  Does it describe the risks of the treatment/exercise?", bold_font)
    write_line("  Q12  Does it describe what happens with no treatment / no exercise?", bold_font)
    write_line("  Q13  Does it describe effects on quality of life?", bold_font)
    write_line("  Q14  Does it make clear there may be more than one option?", bold_font)
    write_line("  Q15  Does it support shared decision-making?", bold_font)
    write_line("  Q16  Overall quality rating (your overall judgment of the publication)", bold_font)
    write_blank()

    write_line("  Scoring guide:  1 = Definite No  |  2 = Mostly No  |  3 = Partially  |  4 = Mostly Yes  |  5 = Definite Yes", normal_font, tip_fill)
    write_blank()

    write_line("STEP 4 — Score JAMA Benchmarks  (columns AI through AL)", bold_font, light_fill)
    write_line("  Mark Y (Yes) or N (No) for each criterion:")
    write_blank()

    write_line("  AI: Authorship — Are the author(s) and their credentials clearly identified?", bold_font)
    write_line("     Social media: Bio states credentials OR verified professional account = Y")
    write_line("  AJ: Attribution — Are references or sources cited?", bold_font)
    write_line("     Social media: Cites studies/sources in caption OR shows references on screen = Y")
    write_line("  AK: Disclosure — Are sponsorships or conflicts of interest declared?", bold_font)
    write_line("     Social media: States sponsorships/partnerships OR uses #ad tags = Y")
    write_line("  AL: Currency — Is a date of creation or update provided?", bold_font)
    write_line("     Social media: Post date is always visible = almost always Y")
    write_blank()

    write_line("STEP 5 — Notes (column AN)", bold_font, light_fill)
    write_line("  Add any notes about the content, especially if you were unsure about a score")
    write_line("  Move to the next row and repeat")
    write_blank()

    # --- AUTO-CALCULATED ---
    write_section("AUTO-CALCULATED COLUMNS (do NOT edit these)")
    write_line("  AE: DISCERN Section 1 subtotal  =  SUM(Q1 through Q8)")
    write_line("  AF: DISCERN Section 2 subtotal  =  SUM(Q9 through Q15)")
    write_line("  AG: DISCERN Total  =  SUM(Q1 through Q16)")
    write_line("  AH: DISCERN Category  (Very Poor / Poor / Fair / Good / Excellent)")
    write_line("  AM: JAMA Total  =  count of Y values in AI through AL  (0 to 4)")
    write_blank()

    write_line("  DISCERN Interpretation:  16-26 Very Poor  |  27-38 Poor  |  39-50 Fair  |  51-62 Good  |  63-80 Excellent", normal_font, tip_fill)
    write_blank()

    # --- BATCHING ---
    write_section("RECOMMENDED BATCHING STRATEGY")
    write_line("  Batch 1:  WEB-01 to WEB-50   (websites)       ~2-3 hours   — fastest, start here", bold_font)
    write_line("  Batch 2:  YT-01  to YT-50    (YouTube)        ~3-4 hours   — need to watch videos")
    write_line("  Batch 3:  TT-01  to TT-50    (TikTok)         ~1.5-2 hours — short videos")
    write_line("  Batch 4:  IG-01  to IG-50    (Instagram)      ~1.5-2 hours — short content")
    write_blank()
    write_line("  Total estimated time: 8-11 hours of focused work", bold_font)
    write_blank()

    # --- INTER-RATER ---
    write_section("INTER-RATER RELIABILITY")
    write_line("  After you finish all 200 items, give the 'Second Rater' sheet to your second rater")
    write_line("  They score the same 40 items (10 per platform) independently")
    write_line("  Do NOT show them your scores beforehand")
    write_line("  Target: ICC >= 0.75 for DISCERN, Cohen's kappa >= 0.61 for JAMA")
    write_blank()

    # --- TIPS ---
    write_section("TIPS FOR CONSISTENT SCORING")
    write_line("  Score one full platform before moving to the next (builds consistency)", tip_font, tip_fill)
    write_line("  Take breaks between batches — scoring fatigue affects quality", tip_font, tip_fill)
    write_line("  If unsure on a score, write a note in column AN and come back later", tip_font, tip_fill)
    write_line("  When in doubt, default to 3 (Partially) rather than guessing high or low", tip_font, tip_fill)
    write_line("  Re-read the DISCERN quick reference card before each new batch", tip_font, tip_fill)
    write_blank()

    # --- AFTER SCORING ---
    write_section("AFTER SCORING IS COMPLETE")
    write_line("  1. Save the spreadsheet")
    write_line("  2. Export 'Data Collection' sheet as CSV (for R analysis)")
    write_line("  3. Run the R analysis script:  analysis/analysis_script.R")
    write_line("  4. The R script will generate all statistics, figures, and tables for Chapter 3")
    write_blank()
    write_line("  When you're done scoring, come back to Claude and say 'run the analysis'", bold_font)

    # Save
    wb.save(SPREADSHEET_PATH)
    print(f"Added 'Scoring Guide' sheet as first tab in {SPREADSHEET_PATH}")
    print(f"Sheet tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
