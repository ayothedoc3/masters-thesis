"""
Compile all platform score CSVs into the master spreadsheet.
Updates both 'Data Collection' (200 items) and 'Second Rater' (40 items) sheets.
"""

import openpyxl
import csv
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET_PATH = os.path.join(BASE_DIR, "data_collection_spreadsheet.xlsx")
SCORES_DIR = os.path.join(BASE_DIR, "scores")

SCORE_FILES = [
    os.path.join(SCORES_DIR, "web_scores.csv"),
    os.path.join(SCORES_DIR, "yt_scores.csv"),
    os.path.join(SCORES_DIR, "tt_scores.csv"),
    os.path.join(SCORES_DIR, "ig_scores.csv"),
]

# Spreadsheet column mapping
# G(7) = Creator_Credentials
# M(13) = Creator_Type
# O(15)-AD(30) = DISCERN Q1-Q16
# AI(35) = JAMA_Authorship
# AJ(36) = JAMA_Attribution
# AK(37) = JAMA_Disclosure
# AL(38) = JAMA_Currency
# AN(40) = Notes

DISCERN_Q_START_COL = 15  # Column O = Q1
JAMA_START_COL = 35       # Column AI = Authorship


def read_all_scores():
    """Read all score CSVs and return dict keyed by Item_ID."""
    all_scores = {}
    for filepath in SCORE_FILES:
        if not os.path.exists(filepath):
            print(f"  WARNING: {filepath} not found, skipping")
            continue
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            count = 0
            for row in reader:
                item_id = row.get("Item_ID", "").strip()
                if item_id:
                    all_scores[item_id] = row
                    count += 1
            print(f"  Read {count} scores from {os.path.basename(filepath)}")
    return all_scores


def apply_scores_to_sheet(ws, scores, sheet_name):
    """Apply scores to a worksheet."""
    updated = 0
    skipped = 0

    for row_num in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row_num, column=1).value
        if not item_id:
            continue

        score_row = scores.get(str(item_id).strip())
        if not score_row:
            print(f"    WARNING: {item_id} not found in scores")
            skipped += 1
            continue

        # Creator_Credentials (col G = 7)
        cred = score_row.get("Creator_Credentials", "").strip()
        if cred:
            ws.cell(row=row_num, column=7, value=cred)

        # Creator_Type (col M = 13)
        ctype = score_row.get("Creator_Type", "").strip()
        if ctype:
            ws.cell(row=row_num, column=13, value=ctype)

        # DISCERN Q1-Q16 (cols O=15 through AD=30)
        for q in range(1, 17):
            q_key = f"Q{q}"
            q_val = score_row.get(q_key, "").strip()
            if q_val:
                try:
                    ws.cell(row=row_num, column=DISCERN_Q_START_COL + q - 1, value=int(q_val))
                except ValueError:
                    # Try float then int
                    try:
                        ws.cell(row=row_num, column=DISCERN_Q_START_COL + q - 1, value=int(float(q_val)))
                    except ValueError:
                        print(f"    WARNING: {item_id} Q{q} invalid value: {q_val}")

        # JAMA criteria (cols AI=35 through AL=38)
        jama_fields = [
            ("JAMA_Authorship", 35),
            ("JAMA_Attribution", 36),
            ("JAMA_Disclosure", 37),
            ("JAMA_Currency", 38),
        ]
        for field_name, col_idx in jama_fields:
            val = score_row.get(field_name, "").strip().upper()
            if val in ("Y", "N"):
                ws.cell(row=row_num, column=col_idx, value=val)
            elif val in ("YES", "TRUE", "1"):
                ws.cell(row=row_num, column=col_idx, value="Y")
            elif val in ("NO", "FALSE", "0"):
                ws.cell(row=row_num, column=col_idx, value="N")

        # Notes (col AN = 40)
        notes = score_row.get("Notes", "").strip()
        if notes:
            ws.cell(row=row_num, column=40, value=notes)

        updated += 1

    print(f"  {sheet_name}: {updated} items updated, {skipped} skipped")
    return updated


def verify_sheet(ws, sheet_name):
    """Quick verification of scored data."""
    filled_discern = 0
    filled_jama = 0
    total = 0

    for row_num in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row_num, column=1).value
        if not item_id:
            continue
        total += 1

        # Check if Q1 (col O=15) has a value
        q1 = ws.cell(row=row_num, column=15).value
        if q1 is not None:
            filled_discern += 1

        # Check if JAMA_Authorship (col AI=35) has a value
        jama_auth = ws.cell(row=row_num, column=35).value
        if jama_auth is not None:
            filled_jama += 1

    print(f"  {sheet_name} verification:")
    print(f"    Total rows: {total}")
    print(f"    DISCERN scored: {filled_discern}/{total}")
    print(f"    JAMA scored: {filled_jama}/{total}")

    # Spot check a few items
    spot_checks = [2, 52, 102, 152]  # WEB-01, YT-01, TT-01, IG-01
    for r in spot_checks:
        if r <= ws.max_row:
            item_id = ws.cell(row=r, column=1).value
            q1 = ws.cell(row=r, column=15).value
            q16 = ws.cell(row=r, column=30).value
            total_formula = ws.cell(row=r, column=33).value
            cat_formula = ws.cell(row=r, column=34).value
            jama_auth = ws.cell(row=r, column=35).value
            creator_type = ws.cell(row=r, column=13).value
            print(f"    {item_id}: Q1={q1}, Q16={q16}, Creator={creator_type}, JAMA_Auth={jama_auth}")


def main():
    print("=" * 60)
    print("Compiling scores into spreadsheet")
    print("=" * 60)
    print()

    # Read all scores
    print("Reading score files...")
    scores = read_all_scores()
    print(f"  Total scores loaded: {len(scores)}")
    print()

    if len(scores) == 0:
        print("ERROR: No scores found. Check that score CSV files exist in data/scores/")
        return

    # Open spreadsheet
    wb = openpyxl.load_workbook(SPREADSHEET_PATH)

    # Update Data Collection sheet
    print("Updating 'Data Collection' sheet...")
    ws_data = wb["Data Collection"]
    apply_scores_to_sheet(ws_data, scores, "Data Collection")
    print()

    # Update Second Rater sheet
    print("Updating 'Second Rater' sheet...")
    ws_rater = wb["Second Rater"]
    apply_scores_to_sheet(ws_rater, scores, "Second Rater")
    print()

    # Save
    print("Saving spreadsheet...")
    wb.save(SPREADSHEET_PATH)
    print(f"  Saved to: {SPREADSHEET_PATH}")
    print()

    # Verify
    print("Verifying...")
    wb2 = openpyxl.load_workbook(SPREADSHEET_PATH)
    verify_sheet(wb2["Data Collection"], "Data Collection")
    print()
    verify_sheet(wb2["Second Rater"], "Second Rater")

    print()
    print("=" * 60)
    print("SUCCESS: All scores compiled into spreadsheet!")
    print("=" * 60)


if __name__ == "__main__":
    main()
