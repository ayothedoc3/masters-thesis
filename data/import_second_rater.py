"""
Import second rater scores into the Second Rater sheet of the spreadsheet,
then export to CSV for R/Python analysis.
"""
import openpyxl
import csv
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET_PATH = os.path.join(BASE_DIR, "data_collection_spreadsheet.xlsx")
SCORES_PATH = os.path.join(BASE_DIR, "scores", "second_rater_scores.csv")
OUTPUT_CSV = os.path.join(BASE_DIR, "csv", "second_rater_scores.csv")

DISCERN_Q_START_COL = 15
JAMA_START_COL = 35


def main():
    # Read second rater scores
    scores = {}
    with open(SCORES_PATH, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            scores[row["Item_ID"].strip()] = row
    print(f"Read {len(scores)} second rater scores")

    # Update spreadsheet
    wb = openpyxl.load_workbook(SPREADSHEET_PATH)
    ws = wb["Second Rater"]
    updated = 0

    for row_num in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row_num, column=1).value
        if not item_id:
            continue
        score_row = scores.get(str(item_id).strip())
        if not score_row:
            print(f"  WARNING: {item_id} not found in second rater scores")
            continue

        # Creator_Credentials (col G=7)
        cred = score_row.get("Creator_Credentials", "").strip()
        if cred:
            ws.cell(row=row_num, column=7, value=cred)

        # Creator_Type (col M=13)
        ctype = score_row.get("Creator_Type", "").strip()
        if ctype:
            ws.cell(row=row_num, column=13, value=ctype)

        # DISCERN Q1-Q16
        for q in range(1, 17):
            val = score_row.get(f"Q{q}", "").strip()
            if val:
                try:
                    ws.cell(row=row_num, column=DISCERN_Q_START_COL + q - 1, value=int(val))
                except ValueError:
                    pass

        # JAMA
        for field, col in [("JAMA_Authorship", 35), ("JAMA_Attribution", 36),
                           ("JAMA_Disclosure", 37), ("JAMA_Currency", 38)]:
            val = score_row.get(field, "").strip().upper()
            if val in ("Y", "N"):
                ws.cell(row=row_num, column=col, value=val)

        # Notes
        notes = score_row.get("Notes", "").strip()
        if notes:
            ws.cell(row=row_num, column=40, value=notes)

        updated += 1

    wb.save(SPREADSHEET_PATH)
    print(f"Updated {updated} rows in Second Rater sheet")

    # Now export Second Rater sheet to CSV with computed values
    wb2 = openpyxl.load_workbook(SPREADSHEET_PATH)
    ws2 = wb2["Second Rater"]

    COLUMNS = {
        1: "Item_ID", 2: "Platform", 3: "Search_Term", 4: "URL",
        5: "Title_Caption", 6: "Creator_Name", 7: "Creator_Credentials",
        8: "Date_Posted", 9: "Views", 10: "Likes", 11: "Shares",
        12: "Comments", 13: "Creator_Type", 14: "Content_Format",
    }
    for q in range(1, 17):
        COLUMNS[14 + q] = f"DISCERN_Q{q}"
    COLUMNS[31] = "DISCERN_Section1"
    COLUMNS[32] = "DISCERN_Section2"
    COLUMNS[33] = "DISCERN_Total"
    COLUMNS[34] = "DISCERN_Category"
    COLUMNS[35] = "JAMA_Authorship"
    COLUMNS[36] = "JAMA_Attribution"
    COLUMNS[37] = "JAMA_Disclosure"
    COLUMNS[38] = "JAMA_Currency"
    COLUMNS[39] = "JAMA_Total"
    COLUMNS[40] = "Notes"

    headers = [COLUMNS[i] for i in sorted(COLUMNS.keys())]
    rows_written = 0

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for row_num in range(2, ws2.max_row + 1):
            item_id = ws2.cell(row=row_num, column=1).value
            if not item_id:
                continue
            row_data = []
            for col_idx in sorted(COLUMNS.keys()):
                cell = ws2.cell(row=row_num, column=col_idx)
                value = cell.value
                col_name = COLUMNS[col_idx]

                if isinstance(value, str) and value.startswith("="):
                    if col_name == "DISCERN_Section1":
                        vals = [ws2.cell(row=row_num, column=c).value for c in range(15, 23)]
                        value = sum(v for v in vals if isinstance(v, (int, float)))
                    elif col_name == "DISCERN_Section2":
                        vals = [ws2.cell(row=row_num, column=c).value for c in range(23, 30)]
                        value = sum(v for v in vals if isinstance(v, (int, float)))
                    elif col_name == "DISCERN_Total":
                        vals = [ws2.cell(row=row_num, column=c).value for c in range(15, 31)]
                        value = sum(v for v in vals if isinstance(v, (int, float)))
                    elif col_name == "DISCERN_Category":
                        total = sum(v for v in [ws2.cell(row=row_num, column=c).value for c in range(15, 31)] if isinstance(v, (int, float)))
                        if total <= 26: value = "Very Poor"
                        elif total <= 38: value = "Poor"
                        elif total <= 50: value = "Fair"
                        elif total <= 62: value = "Good"
                        else: value = "Excellent"
                    elif col_name == "JAMA_Total":
                        jvals = [ws2.cell(row=row_num, column=c).value for c in range(35, 39)]
                        value = sum(1 for v in jvals if v == "Y")

                row_data.append(value if value is not None else "")
            writer.writerow(row_data)
            rows_written += 1

    print(f"Exported {rows_written} rows to {OUTPUT_CSV}")

    # Verify variation
    r1_scores = {}
    with open(os.path.join(BASE_DIR, "csv", "master_dataset.csv"), "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row["Item_ID"] in scores:
                r1_scores[row["Item_ID"]] = int(float(row["DISCERN_Total"])) if row["DISCERN_Total"] else 0

    r2_scores_check = {}
    with open(OUTPUT_CSV, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            r2_scores_check[row["Item_ID"]] = int(float(row["DISCERN_Total"])) if row["DISCERN_Total"] else 0

    print("\nScore comparison (Rater 1 vs Rater 2):")
    diffs = []
    for iid in sorted(r1_scores.keys()):
        r1 = r1_scores.get(iid, 0)
        r2 = r2_scores_check.get(iid, 0)
        diff = abs(r1 - r2)
        diffs.append(diff)
        marker = " ***" if diff > 5 else ""
        print(f"  {iid}: R1={r1}, R2={r2}, diff={diff}{marker}")

    if diffs:
        print(f"\n  Mean absolute difference: {sum(diffs)/len(diffs):.1f}")
        print(f"  Max difference: {max(diffs)}")
        print(f"  Items with diff > 5: {sum(1 for d in diffs if d > 5)}")


if __name__ == "__main__":
    main()
